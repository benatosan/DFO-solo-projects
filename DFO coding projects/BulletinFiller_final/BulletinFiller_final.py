
'''
##Automating GLLevels and GLForecast data entry to excel
##Fills table for specific lastMonth in BackPageTable.xlsx and Great Lakes Forecast yyyy.xlsx
##User still needs to copy over to For Publisher Sheet
##Maybe future could completely autommate, don't wanna mess up links to publisher
##note: Apologies for the varying syntax and tendencies, this project spanned a long time.

For any issues contact khaleel.arfeen@dfo-mpo.gc.ca
'''

##importing libraries
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook, Workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta
from zipfile import BadZipFile
from text_unidecode import unidecode
import xlwings as xw

import tkinter as tk
import sys
import threading
import warnings


class bulletin:
    '''
    Parent class for tkinter GUI
    holds all functions for data entry
    '''

    #Current year and month for directory nav
    year = datetime.now().year   
    month = datetime.now().strftime('%m')    #gives month w/ leading 0
    
    currMonth = datetime.now()
    prevMonth = currMonth - relativedelta(months=1)
    lastMonth = prevMonth.strftime('%m')      #for files named for a previous month

    month_dict = {
    '01': 'January',
    '02': 'February',
    '03': 'March',
    '04': 'April',
    '05': 'May',
    '06': 'June',
    '07': 'July',
    '08': 'August',
    '09': 'September',
    '10': 'October',
    '11': 'November',
    '12': 'December'
    }

    lakeList = [
        'Superior',
        'Huron',
        'Clair',
        'Erie',
        'Ontario',
        'Montreal'
    ]

    #Tab warnings fill cmd prompt but causes no issues so ignore
    warnings.filterwarnings("ignore")

    def __init__(self):
        '''
        Set instance vars
        '''
        #Filenames for the forecast files
        self.levelFile = f"GLLevels{self.year}{self.month}.txt"
        self.forecastFile = f"GLForecast{self.year}{self.month}.txt"

        if self.month == '01':
            self.lastMonth = '12'
            self.year -= 1
        
    #--------------------------------------------------------------------------------------------------

    def get_levels(self) -> str:
        """
        Returns the inputted file from forecast files directory as a csv.
        """
        #Open data files
        try:
            f = pd.read_csv(Path(r'I:\Tides\BULLETIN\Forecast Files') / str(self.year) / self.levelFile, sep="\\s{3,}", header=None, skiprows = 8, engine='python')
        except FileNotFoundError:
            print("GLLevels file not found, check I:\\Tides\\BULLETIN\\Forecast Files\\" + str(self.year) + "\\GLLevels" + str(self.year) + str(self.month) + ".txt")
            return
           
        f = f[f[0] != "at Jetty No.1"]      #takes out unnecessary row for MTL
        return f
    
    #--------------------------------------------------------------------------------------------------

    #This function is technically functional but openpyxl doesn't support xlsm and therefore the file gets corrupted
    #Cannot use until xlsm support or change to using xlwings or alike
    def update_network(self) -> None:
    
        #Updates the Network Means yyyy xlsm file

        #Make dictionary that relates lake name and month's water level
        f = self.get_levels()   #read in GLLevelsyyyymm as csv
        lakeData = {}
        for i in range(0, 6):   #Iterate through GLLevels
            lakeName = f.iloc[i,0]  
            for lake in self.lakeList:  #Error catching if lake spelling doesn't match program needs
                if lake in lakeName:
                    lakeName = lake
                    break
            lakeLevel = f.iloc[i,3]
            lakeData[lakeName] = lakeLevel  #Connect lake's name to the month's water level of the lake

        #Describe excel file to write to
        fileName = "Network Means " + str(self.year) + ".xlsm"
        file = Path(r'I:\Tides\Great Lakes Network Means') / str(self.year) / fileName

        #Opening excel file
        try:
            book = xw.Book(file)
        except FileNotFoundError:
            print(f"file not found, Network Means {str(self.year)}.xlsm file should be in I:\\Tides\\Great Lakes Network Mean\\{str(self.year)}")
            return
        except BadZipFile:
            print(f"Error: The file '{fileName}' is not a valid Excel file or is corrupted. Please fix table from backup folder.")
            return # Exit the function early
        
        #Initialize vars to avoid unbound error
        myRow = None
        myCol = None

        #User progress notice
        print("Updating Network Means ", self.year, "...")

        #Enters data to the excel file
        for sheet in book.sheets:   #Goes through each sheet
            if "_Means" not in sheet.name:
                continue
            for key in lakeData:    #Goes through the dictionary
                if key in sheet.name:    #if lake's name is in the sheet's name
                    for row in range(3, sheet.range('A2').end('down').row + 1):
                        if str(int(sheet.range(f'A{row}').value)) == str(self.year):   #Finds this year's row
                            myRow = str(row)
                            break
                    for col in range(2, sheet.range('B2').end('right').column + 1):
                        if sheet.range(f'{chr(64+col)}2').value == self.month_dict[self.lastMonth]:   #Finds last month's column (enumerate returns tuple hence colDat)
                            myCol = chr(64+col)
                            break
                    if myCol != None and myRow != None: #Error catch if col or row could not find the correct cell
                        sheet.range(myCol + myRow).value = lakeData[key]    #Enters corresponding lake's data into the found cell
                        break
                    else:
                        print("Error: myCol or myRow empty") 
                        return
        print("Saving...")
        book.save()
        print(f"Network Means {self.year} updated")
    #--------------------------------------------------------------------------------------------------
    
    def populate_BackPage(self, sheetName: str) -> None:
        """
        Populates excel file. To be used with BackPageTable.xlsx only. Sensitive to formatting of excel.
        """

        f = self.get_levels()
        
        #Initializing and opening excel file
        fileName = "BackPageTable.xlsx"
        file = Path(r'I:\Tides\BULLETIN\Production\Publisher') / fileName
        try:
            book = load_workbook(file)
        except FileNotFoundError:
            print("file not found, BackPageTable.xlsx should be in I:\\Tides\\BULLETIN\\Production\\Publisher")
            return
        except BadZipFile:
            print(f"Error: The file '{fileName}' is not a valid Excel file or is corrupted. Please fix table from backup folder.")
            return # Exit the function early
        
        #Checks that sheet exists     
        if sheetName in book.sheetnames:
            sheet = book[sheetName]
            print("Found sheet")
    
            #Data entry into correct cells
            for i in range(0, 6):
                sheet[chr(66+i)+'4'] = f.iloc[i,1]
                sheet[chr(66+i)+'3'] = f.iloc[i,3]
                sheet[chr(66+i)+'10'] = f.iloc[i,4]
            book.save(file)
        else:
            print("current month sheet not found in excel file, check I:\\Tides\\BULLETIN\\Production\\Publisher\\BackPageTable.xlsx and ensure no spaces and proper capitalization (ex: 'Oct')")
            return
        print("BackPageTable filled for month of "+ sheetName + "! :)") 

    #--------------------------------------------------------------------------------------------------

    #--------------------------------------------------------------------------------------------------
    #Automating forecast file data entry to Great Lakes Forecast yyyy.xlsx
    #--------------------------------------------------------------------------------------------------

    def get_forecast(self) -> str:
        '''
        Returns the forecast file from the directory.
        '''
        return Path(r'I:\Tides\BULLETIN\Forecast Files') / str(self.year) / self.forecastFile
    
    #--------------------------------------------------------------------------------------------------
    
    def get_chunks(self) -> dict:
        '''
        Compiles a list of data separated into each lake and each month
        and makes a dictionary relating lake names to the data using make_dict() function

        Dictionary will look like:
            'Lake name' : [[month's data], [next month's data], [etc.]],
            'Next lake name' : [[month's data], [next month's data], [etc.]]

        The name of the month will be in month's data[0]
        '''
        #List holding each Lake's complete data
        LakeChunks = []
        
        #List for lake name, matches index of LakeChunks (Lake for LakeChunks[0] = Lakes[0])
        Lakes = []
        try:
            #Opening forecast file and skipping descriptive lines
            with open(self.get_forecast(), 'r') as file:
                chunk = []
                monthData = []
                for skip in range(0,8):
                    next(file)
        
                #Reads line by line and makes a list holding only the desired values
                for badline in file:
                    if badline.strip():     #"badline" is a line with all the whitespace
                        line = badline.strip().split('    ')    #split it into columns as shown on the txt file
                        #Makes a list holding lake names
                        if 'LAKE' in line[0] or 'MONTREAL' in line[0]:
                            if Lakes:      #false when Lakes empty (first time through)
                                LakeChunks.append(chunk)    #appends chunk before the current Lake
                                chunk = []
                            Lakename = line[0].split('(')   #split before chart datum and igld values
                            Lakes.append(Lakename[0])   #Only store lake name

                        #For non-header lines make a chunk of data for that lake
                        else:
                            for ele in line:    #loop through a single line in txt file
                                if ele != '':   #Ignore whitespace
                                    monthData.append(ele.strip())   #Clean up whitespace around values
                            chunk.append(monthData)     #Add month's data to that lake's data list
                            monthData = []
            
                #Catches last chunk
                LakeChunks.append(chunk)
                chunk = []
        
                return self.make_dict(LakeChunks, Lakes)
        except FileNotFoundError:
            print(f"GLForecast file not found, check that the file {str(self.forecastFile)} exists in I:\Tides\BULLETIN\Forecast Files")
            return
    
    #--------------------------------------------------------------------------------------------------
    
    def make_dict(self, lakechunks: list, lakenames: list) -> dict:
        '''
        Creates a dictionary with lakenames as key and lakechunks as value
        '''
        myDict = {}
        for i in range(0,len(lakenames)):
            myDict[lakenames[i]] = lakechunks[i]
    
        return myDict
    
    #--------------------------------------------------------------------------------------------------
    
    def populate_GLF(self, data: dict, sheetName: str) -> None:
        '''
        Populates the Great Lakes Forecast excel file
        Again, sensitive to formatting of excel and txt file
        '''
    
        #Opening excel file
        fileName = f"Great Lakes Forecast {self.year}.xlsx"
        file = Path(r'I:\Tides\BULLETIN\Excel Files') / str(self.year) / fileName
        try:
            book = load_workbook(file)
        except FileNotFoundError:
            print(r"file not found, make sure it exists in I:\Tides\BULLETIN\Excel Files\yyyy under " + fileName)
            return
        except BadZipFile:
            print(f"Error: The file '{fileName}' is not a valid Excel file or is corrupted. Please recreate the file from the xls in archive and remove spaces after month names.")
            return # Exit the function early   
    
        #Accessing sheet
        if sheetName in book.sheetnames:
            sheet = book[sheetName]

            #Going through excel file
            for row in range(1, 55):
                if sheet['A'+str(row)].value in data.keys():
                    currKey = sheet['A'+str(row)].value
                    cursor = data[currKey]                               #Takes lake name to decide which lake's data to use
                    count = 0         #counts rows within a lake chunk
                    countCell = 0     #Iterates through cells
        
                #Enters both value and parentheses value into correct cell
                elif sheet['A'+str(row)].value == cursor[count][0]:
                    for col in range(1, len(cursor[count]), 2):
                        sheet[chr(66+countCell)+str(row)] = cursor[count][col] + ' ' + cursor[count][col+1]
                        countCell+=1
                    count += 1
                    if count >= len(cursor):
                        count = 0
                    countCell = 0
        
                #For cases where data is not needed
                else:
                    pass
        else:
            print(f"Sheet not found, check {fileName} in {str(file)}")
            quit()
    
        book.save(file)
        print("Great Lakes Forecast filled for month of "+ sheetName + "! :)")

    #--------------------------------------------------------------------------------------------------

    def populate_graphs(self) -> None:
        '''
        Creates an excel file, "Temporary_Bulletin_Graph_Data.xlsx" matching the data which must be entered into BulletinGraphs.xlsx
        User must copy and paste the data over to BulletinGraphs.xlsx
        '''
        #Load in data
        data = self.get_chunks()
        newData = {}     #will use this since dict is immutable
        chartDatum  = {}    #holds chart datums
        #Make dict where keys in data match lakelist names
        # ex: LAKE ST. CLAIR -> Clair
        for longName in data:   #Iterates through raw lake names from GLForecastyyyymm.txt
            for lakeName in self.lakeList:  #Iterates through names from lakelist
                if lakeName.lower() in longName.lower():    #if they are a match
                    newData[lakeName] = data[longName]  #New dict with key from lakelist gets associated data
        #Directory for excel file
        fileName = "BulletinGraphs.xlsx"
        file = Path(r'I:\Tides\BULLETIN\Production\Publisher') / fileName

        #Opening excel file to read in chart datum
        try:
            book = load_workbook(file)
        except FileNotFoundError:
            print(f"file not found, {fileName} should be found in {str(file)}")
            return
        except BadZipFile:
            print(f"Error: The file '{fileName}' is not a valid Excel file or is corrupted. Please fix table from backup folder.")
            return # Exit the function early

        for k, sheetName in enumerate(book.sheetnames, 0):   #Iterates through sheets
            if "Data" not in sheetName: #All sheets of interest have Data in the name
                continue    #skip iteration for efficiency
            
            for key in newData: #Iterate through lake names
                if unidecode(key) in unidecode(sheetName):    #if sheet is for lake of interest. Also, if can't find a true case, sheet isn't defined and function will error out.
                    sheet = book[sheetName] #set sheetname as active sheet
                    if sheet['A41'].value == "CHART DATUM" and sheet['C41'].value != None:  #set chartdatum var for this lake
                        chartDatum[key] = sheet['C41'].value
                        if key in self.get_levels().iloc[k-1, 0]:   #ensure chart datum is for correct lake
                            chartDatum[key] = round(self.get_levels().iloc[k-1, 3] - chartDatum[key], 2)    #chartDatum holds the current month data to be entered into excel file
                        else:
                            print("Bulletin graphs lake order not same as GLLevels, please rearrange accordingly")
                            return
                    else:
                        print("Error: make sure CHART DATUM is in cell A41 and the value is in cell C41")
                        return
                    break
                    

        #Create a temporary workbook to write data into
        writeBook = Workbook()
        sheet = writeBook.active

        #initialize local variables
        col = 1
        row = 1

        for lake, lakeData in newData.items():  #Iterate through dict assigning lake names to their data
            sheet.cell(row = row, column=1, value=lake) #First cell for each lake is lake name
            row+=1
            sheet.cell(row = row, column=1, value= self.month_dict[self.lastMonth]) #write in the month so user knows where to put data in
            for i in range(2,5):    #First line has Water level minus chart datum value
                sheet.cell(row = row, column=i, value=chartDatum[lake]) #Enter value 3 times
            row+=1
            for monthData in lakeData:  #Go through each month's data for the lake
                sheet.cell(row = row, column= col, value = monthData[0])    #Month name
                sheet.cell(row = row, column= col+1, value = float(monthData[2].replace('(', '').replace(')', ''))) #Forecast Range - High
                sheet.cell(row = row, column= col+2, value = float(monthData[6].replace('(', '').replace(')', ''))) #Forecast Range - Low
                row+=1
            row+=1  #Skip line before next lake for user readability

        #Define file path for created excel sheet
        writeFilePath = Path(r'I:\Tides\BULLETIN\Production\Publisher\Temporary_Bulletin_Graph_Data.xlsx')
        writeBook.save(writeFilePath)
        print("Temporary Graph data excel file created in ", str(writeFilePath))    #Inform user where to find file
        return
    #--------------------------------------------------------------------------------------------------

    #--------------------------------------------------------------------------------------------------
    #Running functions
    #--------------------------------------------------------------------------------------------------

    def run(self, sheetName: str) -> None:
        '''
        Runs above functions to fill excel files
        '''
        #Enter data to BackPageTable.xlsx
        self.populate_BackPage(sheetName)

        #Enter data to Great Lakes Forecast yyyy.xlsx
        WLFCdata = self.get_chunks()
        self.populate_GLF(WLFCdata, sheetName)

        #Make excel file holding all Bulletin Graphs data for easy copy-paste
        self.populate_graphs()
        
        #Update network means yyyy excel file
        self.update_network()
    
    #--------------------------------------------------------------------------------------------------
    
    def threaded_run(self, sheetName: str) -> None:
        '''
        run() is too long for tkinter, start run on diff thread so root.mainloop() can go
        '''
        def run_in_thread():
            '''
            Runs a run() thread and prints final msg in mainloop thread since tkinter not thread safe.
            '''
            self.run(sheetName)
            #final msg to tkinter in mainloop thread to avoid runtime error
            self.root.after(0, lambda: print("You can now close this window"))
        threading.Thread(target=run_in_thread).start()
    

#tkinter GUI
class tkintBulletin(bulletin):
    '''
    tkinter GUI class
    '''
    class IO_redirector(object):
        #Sets its own textArea
        def __init__(self, textArea):
            self.textArea = textArea
    
    class stdout_redirect(IO_redirector):
        #Inits IO_redirector and writes to textArea
        def write(self, str):
            self.textArea.after(0, self.textArea.insert, tk.END, str)

    #--------------------------------------------------------------------------------------------------

    def __init__(self):
        #Initializing vars and making instance of parent class (bulletin)
        super().__init__()
        self.monthList = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        self.root = tk.Tk()
        self.name = tk.StringVar()
    
    #--------------------------------------------------------------------------------------------------

    def set_and_print(self, m:str) -> None:
        '''
        For button in GUI, prints and sets selection for clarity
        '''
        self.name.set(m)
        print("Will write to BackPageTable.xlsx and Great Lakes Forecast " + str(self.year) + ".xlsx for the month of " + self.name.get() + " and create an excel file for BulletinGraphs")

    #--------------------------------------------------------------------------------------------------

    def on_close(self):
        '''
        Resets stdout and closes tkinter window
        '''
        sys.stdout = sys.__stdout__
        self.root.destroy()
        sys.exit()

    #--------------------------------------------------------------------------------------------------

    def interface(self):
        '''
        Runs interface for tkinter GUI
        '''
        #Title label
        label = tk.Label(self.root, text = "Current data is from " + self.forecastFile + " and " + self.levelFile + "\nEnsure you choose the correct month below to fill in the excel files.\n\nDirectory for excel files:\n I:\\Tides\\BULLETIN\\Production\\Publisher for BackPageTable\n I:\\Tides\\BULLETIN\\Excel Files\\yyyy for Great Lakes Forecast\n I:\\Tides\\BULLETIN\\Production\\Publisher for Temp Bulletin Graph Data", width = 80, anchor = 'center')
        label.pack()
        
        #Creates button list
        for m in self.monthList:
            tk.Button(self.root, text = m, command=lambda m=m: self.set_and_print(m), width = 10).pack()
        
        #RUN button runs the program
        runButton = tk.Button(self.root, text="RUN", command=lambda: self.threaded_run(self.name.get()))
        runButton.pack()
        
        #defining textbox
        outputBox = tk.Text(self.root, height = 10, width = 60)
        outputBox.pack()
        
        #outputBox is IO_redirector, stdout becomes the textArea so print() -> outputBox
        sys.stdout = tkintBulletin.stdout_redirect(outputBox)
        # Set the window close protocol
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)       
        #Mainloop
        self.root.mainloop()
        
    #--------------------------------------------------------------------------------------------------

#Workflow
tkint_bulletin_instance = tkintBulletin()
tkint_bulletin_instance.interface()